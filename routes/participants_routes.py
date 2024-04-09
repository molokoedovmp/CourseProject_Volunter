from flask import Flask, render_template, request, redirect,send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from sqlalchemy import or_

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
db = SQLAlchemy(app)


class Request(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    surname = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"Request(id={self.id}, name={self.name}, surname={self.surname}, email={self.email})"


@app.route('/')
def index():
    requests = get_all_requests()
    return render_template('index.html', requests=requests)


@app.route('/students')
def students():
    requests = get_all_requests()
    return render_template('students.html', requests=requests)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Обработка логики регистрации
        # ...
        return redirect('/')
    else:
        return render_template('registration.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Обработка логики входа
        # ...
        return 'Login Successful'  # Пример ответа
    else:
        return render_template('login.html')


@app.route('/create', methods=['POST'])
def create():
    name = request.form['name']
    surname = request.form['surname']
    email = request.form['email']
    create_request(name, surname, email)
    return redirect('/')


@app.route('/update/<int:request_id>', methods=['GET', 'POST'])
def update(request_id):
    request_obj = get_request_by_id(request_id)

    if request.method == 'POST':
        name = request.form['name']
        surname = request.form['surname']
        email = request.form['email']
        update_request(request_id, name, surname, email)
        return redirect('/')
    else:
        return render_template('update.html', request=request_obj)


@app.route('/delete/<int:request_id>')
def delete(request_id):
    delete_request(request_id)
    return redirect('/')


@app.route('/search', methods=['GET', 'POST'])
def search():
    if request.method == 'POST':
        search_query = request.form['search_query']
        requests = search_requests(search_query)
        return render_template('index.html', requests=requests)
    else:
        return redirect('/')


@app.route('/export', methods=['GET'])
def export():
    filename = export_requests_to_excel()
    return send_file(filename, as_attachment=True)


def create_request(name, surname, email):
    request_obj = Request(name=name, surname=surname, email=email)
    db.session.add(request_obj)
    db.session.commit()


def get_all_requests():
    return Request.query.all()


def get_request_by_id(request_id):
    return Request.query.get(request_id)


def update_request(request_id, name, surname, email):
    request_obj = Request.query.get(request_id)
    request_obj.name = name
    request_obj.surname = surname
    request_obj.email = email
    db.session.commit()


def delete_request(request_id):
    request_obj = Request.query.get(request_id)
    db.session.delete(request_obj)
    db.session.commit()


def search_requests(search_query):
    return Request.query.filter(
        or_(
            Request.name.contains(search_query),
            Request.surname.contains(search_query),
            Request.email.contains(search_query)
        )
    ).all()


def export_requests_to_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Requests"

    requests = get_all_requests()

    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Surname'
    worksheet['D1'] = 'Email'

    row = 2
    for request in requests:
        worksheet.cell(row=row, column=1).value = request.id
        worksheet.cell(row=row, column=2).value = request.name
        worksheet.cell(row=row, column=3).value = request.surname
        worksheet.cell(row=row, column=4).value = request.email
        row += 1

    filename = 'requests.xlsx'
    workbook.save(filename)

    return filename


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
