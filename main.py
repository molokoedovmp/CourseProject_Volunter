from flask import Flask, render_template, request, redirect, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from sqlalchemy import or_, func, desc
import datetime
from io import BytesIO
import pandas as pd


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///volunteer3.db'
db = SQLAlchemy(app)

# region Tables
class Request(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    surname = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"Request(id={self.id}, name={self.name}, surname={self.surname}, email={self.email})"


class Student(db.Model):
    __tablename__ = 'students'
    student_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    surname = db.Column(db.String(100), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('courses.course_id'), nullable=False)
    messengers = db.Column(db.String(100), nullable=False)
    volunteers = db.relationship('Volunteer', backref='student', lazy=True)

    def __repr__(self):
        return f"Student(student_id={self.student_id}, name={self.name}, " \
               f"surname={self.surname}, course_id={self.course_id}, messengers={self.messengers})"


class Event(db.Model):
    __tablename__ = 'events'
    event_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    date = db.Column(db.String(100), nullable=False)
    time_start = db.Column(db.String(100), nullable=False)
    time_end = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    skill_id = db.Column(db.Integer, db.ForeignKey('skills.skill_id'), nullable=False)
    address = db.Column(db.String(100), nullable=False)
    status_id = db.Column(db.Integer, db.ForeignKey('statuses.status_id'), nullable=False)
    curator_id = db.Column(db.Integer, db.ForeignKey('curators.curator_id'), nullable=False)
    volunteers = db.relationship('Volunteer', backref='event', lazy=True)

    def __repr__(self):
        return f"Event(event_id={self.event_id}, name={self.name}, date={self.date}, time_start={self.time_start}" \
               f", time_end={self.time_end}, amount={self.amount}, skill_id={self.skill_id}, address={self.address}, " \
               f"status_id={self.status_id}, curator_id={self.curator_id})"

class Volunteer(db.Model):
    __tablename__ = 'volunteers'
    volunteer_id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('events.event_id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('students.student_id'), nullable=False)
    grade = db.Column(db.Integer, nullable=False)
    comment = db.Column(db.String(100), nullable=True)

    def __repr__(self):
        return f"Volunteer(volunteer_id={self.volunteer_id}, event_id={self.event_id}, student_id={self.student_id}, grade={self.grade}, comment={self.comment})"

class Curator(db.Model):
    __tablename__ = 'curators'
    curator_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    surname = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    messengers = db.Column(db.String(100), nullable=False)
    events = db.relationship('Event', backref='curator', lazy=True)

    def __repr__(self):
        return f"Curator(curator_id={self.curator_id}, name={self.name}, surname={self.surname})"

class Skill(db.Model):
    __tablename__ = 'skills'
    skill_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"Skill(skill_id={self.skill_id}, name={self.name}"

class Status(db.Model):
    __tablename__ = 'statuses'
    status_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"Status(status_id={self.status_id}, name={self.name})"

class Course(db.Model):
    __tablename__ = 'courses'
    course_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

    def __repr__(self):
        return f"Course(status_id={self.course_id}, name={self.name})"
# endregion

# region Routes
@app.route('/')
def index():
    requests = Request.query.all()
    return render_template('participants.html', requests=requests)

@app.route('/application')
def application():
    requests = Request.query.all()
    return render_template('application.html', requests=requests)

@app.route('/home')
def home():
    event = Event.query.all()
    return render_template('home.html', event=event)

@app.route('/students')
def students():
    students = Student.query.all()
    courses = Course.query.all()
    return render_template('students.html', students=students, courses=courses)

@app.route('/events')
def events():
    events = Event.query.all()
    curators = Curator.query.all()
    skills = Skill.query.all()
    statuses = Status.query.all()
    return render_template('events.html', events=events, curators=curators, skills=skills, statuses=statuses)

@app.route('/volunteers')
def volunteers():
    volunteers = Volunteer.query.all()
    events = Event.query.all()
    students = Student.query.all()
    return render_template('volunteers.html', volunteers=volunteers, events=events, students=students)

@app.route('/curator')
def curators():
    curators = Curator.query.all()
    return render_template('curators.html', curators=curators)

@app.route('/skill')
def skills():
    skills = Skill.query.all()
    return render_template('skills.html', skills=skills)

@app.route('/status')
def statuses():
    statuses = Status.query.all()
    return render_template('statuses.html', statuses=statuses)

@app.route('/course')
def courses():
    courses = Course.query.all()
    return render_template('courses.html', courses=courses)

@app.route('/requests')
def requests():
    events = Event.query.all()
    curators = Curator.query.all()
    skills = Skill.query.all()
    statuses = Status.query.all()
    return render_template('requests.html', events=events, curators=curators, skills=skills, statuses=statuses)

@app.route('/request1')
def request1():
     # Выполнение запроса о количестве волонтеров
    volunteers = Volunteer.query.all()
    volunteer_count = len(volunteers)

    # Создание книги Excel
    workbook = Workbook()
    sheet = workbook.active

    # Запись результатов в Excel
    sheet['A1'] = 'Количество волонтеров'
    sheet['B1'] = volunteer_count

    # Сохранение книги Excel в памяти
    excel_data = BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)

    # Установка заголовка Content-Disposition для указания имени файла
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=volunteer_count.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Отправка файла Excel в ответе
    return response

@app.route('/request2')
def request2():
    # Выполнение запроса о средней оценке волонтеров
    avg_grade = db.session.query(func.avg(Volunteer.grade)).scalar()

    # Запрос о средней оценке волонтеров с именем, фамилией и средним значением оценки
    volunteers_data = db.session.query(Student.name, Student.surname,
                                       func.avg(Volunteer.grade)).join(Volunteer,
                                        Volunteer.student_id == Student.student_id).group_by(Student.name, Student.surname).all()

    # Создание таблицы данных с именем, фамилией и средним значением оценки волонтеров
    df = pd.DataFrame(volunteers_data, columns=['First Name', 'Last Name', 'Average Grade'])

    # Создание файла Excel и сохранение данных в него
    excel_file = 'volunteers_data.xlsx'
    df.to_excel(excel_file, index=False)

    # Отправка файла Excel в качестве ответа
    return send_file(excel_file, as_attachment=True)

@app.route('/request4')
def request4():
    # Получение текущего месяца
    current_month = datetime.datetime.now().month

    # Выполнение запроса о количестве мероприятий за текущий месяц
    event_count = db.session.query(func.count(Event.event_id)).filter(func.extract('month', Event.date) == current_month).scalar()

    # Создание таблицы данных с количеством мероприятий за текущий месяц
    df = pd.DataFrame({'Event Count': [event_count]})

    # Создание файла Excel и сохранение данных в него
    excel_file = 'event_count.xlsx'
    df.to_excel(excel_file, index=False)

    # Отправка файла Excel в качестве ответа
    return send_file(excel_file, as_attachment=True)

@app.route('/request5')
def request5():
    # Выполнение запроса на выборку мероприятий с университетским статусом
    events = db.session.query(Event).join(Status).filter(Status.name == 'Университетский').all()

    # Создание таблицы данных с выбранными мероприятиями
    data = {
        'Event ID': [],
        'Name': [],
        'Date': [],
        'Time Start': [],
        'Time End': [],
        'Amount': [],
        'Skill ID': [],
        'Address': [],
        'Status ID': [],
        'Curator ID': []
    }

    for event in events:
        data['Event ID'].append(event.event_id)
        data['Name'].append(event.name)
        data['Date'].append(event.date)
        data['Time Start'].append(event.time_start)
        data['Time End'].append(event.time_end)
        data['Amount'].append(event.amount)
        data['Skill ID'].append(event.skill_id)
        data['Address'].append(event.address)
        data['Status ID'].append(event.status_id)
        data['Curator ID'].append(event.curator_id)

    df = pd.DataFrame(data)

    # Создание файла Excel и сохранение данных в него
    excel_file = 'university_events.xlsx'
    df.to_excel(excel_file, index=False)

    # Отправка файла Excel в качестве ответа
    return send_file(excel_file, as_attachment=True)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Обработка логики регистрации
        # ...
        return redirect('/')
    else:
        return render_template('registration.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Handle the login logic
        # ...
        return render_template('requests.html')  # Example response
    else:
        return render_template('login.html')
# endregion

# region Participants

@app.route('/create', methods=['POST'])
def create():
    name = request.form['name']
    surname = request.form['surname']
    email = request.form['email']
    request_obj = Request(name=name, surname=surname, email=email)
    db.session.add(request_obj)
    db.session.commit()
    return redirect('/')

@app.route('/update/<int:request_id>', methods=['GET', 'POST'])
def update(request_id):
    request_obj = Request.query.get(request_id)

    if request.method == 'POST':
        request_obj.name = request.form['name']
        request_obj.surname = request.form['surname']
        request_obj.email = request.form['email']
        db.session.commit()
        return redirect('/')
    else:
        return render_template('update.html', request=request_obj)

@app.route('/delete/<int:request_id>')
def delete(request_id):
    request_obj = Request.query.get(request_id)
    db.session.delete(request_obj)
    db.session.commit()
    return redirect('/')

@app.route('/search', methods=['GET', 'POST'])
def search():
    if request.method == 'POST':
        search_query = request.form['search_query']
        # Ищем объекты базы данных, удовлетворяющие поисковому запросу
        requests = Request.query.filter(
            or_(
                Request.name.contains(search_query),
                Request.surname.contains(search_query),
                Request.email.contains(search_query)
            )
        ).all()
        return render_template('index.html', requests=requests)
    else:
        return redirect('/')

@app.route('/export', methods=['GET'])
def export():
    # Генерируем файл Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Requests"

    requests = Request.query.all()

    # Заполняем таблицу данными
    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Surname'
    worksheet['D1'] = 'Email'

    row = 2
    for request in requests:
        worksheet.cell(row=row, column=1).value = request.id
        worksheet.cell(row=row, column=2).value = request.name
        worksheet.cell(row=row, column=3).value = request.surname
        worksheet.cell(row=row, column=4).value = request.email
        row += 1

    # Сохраняем файл
    filename = 'requests.xlsx'
    workbook.save(filename)

    # Отправляем файл как вложение
    return send_file(filename, as_attachment=True)

# endregion

# region Students
@app.route('/export/students', methods=['GET'])
def export_students():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"

    students = Student.query.all()

    worksheet['A1'] = 'Student ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Surname'
    worksheet['D1'] = 'Course'
    worksheet['E1'] = 'Messengers'

    row = 2
    for student in students:
        worksheet.cell(row=row, column=1).value = student.student_id
        worksheet.cell(row=row, column=2).value = student.name
        worksheet.cell(row=row, column=3).value = student.surname
        worksheet.cell(row=row, column=4).value = student.course_id
        worksheet.cell(row=row, column=5).value = student.messengers
        row += 1

    filename = 'students.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/students', methods=['GET', 'POST'])
def search_students():
    if request.method == 'POST':
        search_query = request.form['search_query']
        students = Student.query.filter(
            or_(
                Student.name.contains(search_query),
                Student.surname.contains(search_query),
                Student.course_id.contains(search_query),
                Student.messengers.contains(search_query)
            )
        ).all()
        return render_template('students.html', students=students)
    else:
        return redirect('/students')

@app.route('/delete/student/<int:student_id>')
def delete_student(student_id):
    student = Student.query.get(student_id)
    db.session.delete(student)
    db.session.commit()
    return redirect('/students')

@app.route('/update/student/<int:student_id>', methods=['GET', 'POST'])
def update_student(student_id):
    student = Student.query.get(student_id)
    courses = Course.query.all()

    if request.method == 'POST':
        student.name = request.form['name']
        student.surname = request.form['surname']
        student.course_id = request.form['course_id']
        student.messengers = request.form['messengers']
        db.session.commit()
        return redirect('/students')
    else:
        return render_template('update_student.html', student=student, courses=courses)

@app.route('/create/student', methods=['POST'])
def create_student():
    name = request.form['name']
    surname = request.form['surname']
    course_id = request.form['course_id']
    messengers = request.form['messengers']
    student = Student(name=name, surname=surname, course_id=course_id, messengers=messengers)
    db.session.add(student)
    db.session.commit()
    return redirect('/students')

# endregion

#region Volunter CRUD
@app.route('/delete/volunteer/<int:volunteer_id>')
def delete_volunteer(volunteer_id):
    volunteer = Volunteer.query.get(volunteer_id)
    db.session.delete(volunteer)
    db.session.commit()
    return redirect('/volunteers')

@app.route('/update/volunteer/<int:volunteer_id>', methods=['GET', 'POST'])
def update_volunteer(volunteer_id):
    volunteer = Volunteer.query.get(volunteer_id)

    if request.method == 'POST':
        volunteer.event_id = request.form['event_id']
        volunteer.student_id = request.form['student_id']
        volunteer.grade = request.form['grade']
        volunteer.comment = request.form['comment']
        db.session.commit()
        return redirect('/volunteers')
    else:
        return render_template('update_volunteer.html', volunteer=volunteer)

@app.route('/create/volunteer', methods=['POST'])
def create_volunteer():
    event_id = request.form['event_id']
    student_id = request.form['student_id']
    grade = request.form['grade']
    comment = request.form['comment']
    volunteer = Volunteer(event_id=event_id, student_id=student_id, grade=grade, comment=comment)
    db.session.add(volunteer)
    db.session.commit()
    return redirect('/volunteers')

@app.route('/export/volunteers', methods=['GET'])
def export_volunteers():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Volunteers"

    volunteers = Volunteer.query.all()

    worksheet['A1'] = 'Volunteer ID'
    worksheet['B1'] = 'Event ID'
    worksheet['C1'] = 'Student ID'
    worksheet['D1'] = 'Grade'
    worksheet['E1'] = 'Comment'

    row = 2
    for volunteer in volunteers:
        worksheet.cell(row=row, column=1).value = volunteer.volunteer_id
        worksheet.cell(row=row, column=2).value = volunteer.event_id
        worksheet.cell(row=row, column=3).value = volunteer.student_id
        worksheet.cell(row=row, column=4).value = volunteer.grade
        worksheet.cell(row=row, column=5).value = volunteer.comment
        row += 1

    filename = 'volunteers.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/volunteers', methods=['GET', 'POST'])
def search_volunteers():
    if request.method == 'POST':
        search_query = request.form['search_query']
        volunteers = Volunteer.query.filter(
            or_(
                Volunteer.event_id.contains(search_query),
                Volunteer.student_id.contains(search_query),
                Volunteer.grade.contains(search_query),
                Volunteer.comment.contains(search_query)
            )
        ).all()
        return render_template('volunteers.html', volunteers=volunteers)
    else:
        return redirect('/volunteers')
# endregion

#region Events CRUD
@app.route('/export/events', methods=['GET'])
def export_events():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Events"

    events = Event.query.all()

    worksheet['A1'] = 'Event ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Date'
    worksheet['D1'] = 'Time Start'
    worksheet['E1'] = 'Time End'
    worksheet['F1'] = 'Amount'
    worksheet['G1'] = 'Skill ID'
    worksheet['H1'] = 'Address'
    worksheet['I1'] = 'Status'
    worksheet['J1'] = 'Curator ID'

    row = 2
    for event in events:
        worksheet.cell(row=row, column=1).value = event.event_id
        worksheet.cell(row=row, column=2).value = event.name
        worksheet.cell(row=row, column=3).value = event.date
        worksheet.cell(row=row, column=4).value = event.time_start
        worksheet.cell(row=row, column=5).value = event.time_end
        worksheet.cell(row=row, column=6).value = event.amount
        worksheet.cell(row=row, column=7).value = event.skill_id
        worksheet.cell(row=row, column=8).value = event.address
        worksheet.cell(row=row, column=9).value = event.status_id
        worksheet.cell(row=row, column=10).value = event.curator_id
        row += 1

    filename = 'events.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/events', methods=['GET', 'POST'])
def search_events():
    if request.method == 'POST':
        search_query = request.form['search_query']
        events = Event.query.filter(
            or_(
                Event.name.contains(search_query),
                Event.address.contains(search_query),
                Event.status_id.contains(search_query),
                Event.skill_id.contains(search_query)
            )
        ).all()
        return render_template('events.html', events=events)
    else:
        return redirect('/events')

@app.route('/delete/event/<int:event_id>')
def delete_event(event_id):
    event = Event.query.get(event_id)
    db.session.delete(event)
    db.session.commit()
    return redirect('/events')

@app.route('/update/event/<int:event_id>', methods=['GET', 'POST'])
def update_event(event_id):
    event = Event.query.get(event_id)

    if request.method == 'POST':
        event.name = request.form['name']
        event.date = request.form['date']
        event.time_start = request.form['time_start']
        event.time_end = request.form['time_end']
        event.amount = request.form['amount']
        event.skill_id = request.form['skill_id']
        event.address = request.form['address']
        event.status_id = request.form['status']
        event.curator_id = request.form['curator_id']
        db.session.commit()
        return redirect('/events')
    else:
        return render_template('update_event.html', event=event)

@app.route('/create/event', methods=['POST'])
def create_event():
    name = request.form['name']
    date = request.form['date']
    time_start = request.form['time_start']
    time_end = request.form['time_end']
    amount = request.form['amount']
    skill_id = request.form['skill_id']
    address = request.form['address']
    status_id = request.form['status_id']
    curator_id = request.form['curator_id']
    event = Event(name=name, date=date, time_start=time_start, time_end=time_end, amount=amount, skill_id=skill_id, address=address, status_id=status_id, curator_id=curator_id)
    db.session.add(event)
    db.session.commit()
    return redirect('/events')
# endregion

# region Curators
@app.route('/export/curators', methods=['GET'])
def export_curators():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Curators"

    curators = Curator.query.all()


    worksheet['A1'] = 'Curator ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Surname'
    worksheet['D1'] = 'Email'
    worksheet['E1'] = 'Messengers'

    row = 2
    for curator in curators:
        worksheet.cell(row=row, column=1).value = curator.curator_id
        worksheet.cell(row=row, column=2).value = curator.name
        worksheet.cell(row=row, column=3).value = curator.surname
        worksheet.cell(row=row, column=4).value = curator.email
        worksheet.cell(row=row, column=5).value = curator.messengers
        row += 1

    filename = 'curators.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/curators', methods=['GET', 'POST'])
def search_curators():
    if request.method == 'POST':
        search_query = request.form['search_query']
        curators = Curator.query.filter(
            or_(
                Curator.name.contains(search_query),
                Curator.surname.contains(search_query),
                Curator.email.contains(search_query),
                Curator.messengers.contains(search_query)
            )
        ).all()
        return render_template('curators.html', curators=curators)
    else:
        return redirect('/curators')

@app.route('/delete/curator/<int:curator_id>')
def delete_curator(curator_id):
    curator = Curator.query.get(curator_id)
    db.session.delete(curator)
    db.session.commit()
    return redirect('/curators')

@app.route('/update/curator/<int:curator_id>', methods=['GET', 'POST'])
def update_curator(curator_id):
    curator = Curator.query.all()
    skill = Skill.query.all()
    status = Status.query.all()

    if request.method == 'POST':
        curator.name = request.form['name']
        curator.surname = request.form['surname']
        curator.email = request.form['email']
        curator.messengers = request.form['messengers']
        db.session.commit()
        return redirect('/curator')
    else:
        return render_template('update_curator.html', curator=curator, skill=skill, status=status)

@app.route('/create/curator', methods=['POST'])
def create_curator():
    name = request.form['name']
    surname = request.form['surname']
    email = request.form['email']
    messengers = request.form['messengers']
    curator = Curator(name=name, surname=surname, email=email, messengers=messengers)
    db.session.add(curator)
    db.session.commit()
    return redirect('/curator')

# endregion

#region Skills
@app.route('/export/skills', methods=['GET'])
def export_skills():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Skills"

    skills = Skill.query.all()

    worksheet['A1'] = 'Skill ID'
    worksheet['B1'] = 'Name'

    row = 2
    for skill in skills:
        worksheet.cell(row=row, column=1).value = skill.skill_id
        worksheet.cell(row=row, column=2).value = skill.name
        row += 1

    filename = 'skills.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/skills', methods=['GET', 'POST'])
def search_skills():
    if request.method == 'POST':
        search_query = request.form['search_query']
        skills = Skill.query.filter(Skill.name.contains(search_query)).all()
        return render_template('skills.html', skills=skills)
    else:
        return redirect('/skill')

@app.route('/delete/skill/<int:skill_id>')
def delete_skill(skill_id):
    skill = Skill.query.get(skill_id)
    db.session.delete(skill)
    db.session.commit()
    return redirect('/skill')

@app.route('/update/skill/<int:skill_id>', methods=['GET', 'POST'])
def update_skill(skill_id):
    skill = Skill.query.get(skill_id)

    if request.method == 'POST':
        skill.name = request.form['name']
        db.session.commit()
        return redirect('/skill')
    else:
        return render_template('update_skill.html', skill=skill)

@app.route('/create/skill', methods=['POST'])
def create_skill():
    name = request.form['name']
    skill = Skill(name=name)
    db.session.add(skill)
    db.session.commit()
    return redirect('/skill')

    def __repr__(self):
        return f"Skill(skill_id={self.skill_id}, name={self.name})"

#endregion C CC

# region Status CRUD
@app.route('/export/statuses', methods=['GET'])
def export_statuses():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Statuses"

    statuses = Status.query.all()

    worksheet['A1'] = 'Status ID'
    worksheet['B1'] = 'Name'

    row = 2
    for status in statuses:
        worksheet.cell(row=row, column=1).value = status.status_id
        worksheet.cell(row=row, column=2).value = status.name
        row += 1

    filename = 'statuses.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/statuses', methods=['POST'])
def search_statuses():
    search_query = request.form['search_query']
    statuses = Status.query.filter(Status.name.contains(search_query)).all()
    return render_template('statuses.html', statuses=statuses)

@app.route('/delete/status/<int:status_id>')
def delete_status(status_id):
    status = Status.query.get(status_id)
    db.session.delete(status)
    db.session.commit()
    return redirect('/status')

@app.route('/update/status/<int:status_id>', methods=['GET', 'POST'])
def update_status(status_id):
    status = Status.query.get(status_id)

    if request.method == 'POST':
        status.name = request.form['name']
        db.session.commit()
        return redirect('/status')
    else:
        return render_template('update_status.html', status=status)

@app.route('/create/status', methods=['POST'])
def create_status():
    name = request.form['name']
    status = Status(name=name)
    db.session.add(status)
    db.session.commit()
    return redirect('/status')
# endregion

# region Course CRUD
@app.route('/export/courses', methods=['GET'])
def export_courses():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Courses"

    courses = Course.query.all()

    worksheet['A1'] = 'Course ID'
    worksheet['B1'] = 'Name'

    row = 2
    for course in courses:
        worksheet.cell(row=row, column=1).value = course.course_id
        worksheet.cell(row=row, column=2).value = course.name
        row += 1

    filename = 'courses.xlsx'
    workbook.save(filename)

    return send_file(filename, as_attachment=True)

@app.route('/search/courses', methods=['POST'])
def search_courses():
    search_query = request.form['search_query']
    courses = Course.query.filter(Course.name.contains(search_query)).all()
    return render_template('courses.html', courses=courses)

@app.route('/delete/course/<int:course_id>')
def delete_course(course_id):
    course = Course.query.get(course_id)
    db.session.delete(course)
    db.session.commit()
    return redirect('/course')

@app.route('/update/course/<int:course_id>', methods=['GET', 'POST'])
def update_course(course_id):
    course = Course.query.get(course_id)

    if request.method == 'POST':
        course.name = request.form['name']
        db.session.commit()
        return redirect('/course')
    else:
        return render_template('update_course.html', course=course)

@app.route('/create/course', methods=['POST'])
def create_course():
    name = request.form['name']
    course = Course(name=name)
    db.session.add(course)
    db.session.commit()
    return redirect('/course')
# endregion

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)

